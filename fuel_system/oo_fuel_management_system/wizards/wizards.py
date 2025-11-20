import base64
import logging
import os
import tempfile
import openpyxl as xl
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from odoo import models, fields, api
import xlsxwriter
from io import BytesIO


_logger = logging.getLogger(__name__)

WET_SUMMARY_COLS = [
    'Date', 'Opening Stock', 'Deliveries', 'Sales', 'Book Stock','Closing Stock', 'Variance Loss/Gain',
    'Cumulative Variance', 'Cumulative Sales', 'Cumulative Percentage (%)'
]

CASH_SUMMARY_COLS = [
    'DATE', 'PMS', 'PMS AMOUNT', 'AGO', 'AGO AMOUNT', 'BIK', 'BIK AMOUNT', 'LTRS', 'TOTAL AMNT', 'LUBES', 'LPG SALES', 'OTHERS',
    'RECEIPTS', 'TOTAL INCOME', 'CREDIT SALES', 'RTT', 'OTHER EXP.', 'PAYMENTS', 'EXP.BANKING', 'ACT BANKING', 'DIFF' 
]

CREDIT_SUMMARY_COLS = [
    'Date', 'Station', 'LPO', 'Vehicle No.', 'Invoice', 'Account Number', 'Account Name', 'Product', 
    'LTRS/PCS', 'Customer Rate', 'Amount'
]

MID_FONT = Font(name='Arial', bold=True, size=11)
NORMAL_FONT = Font(name='Arial', size=10)


class FmsAnalysis(models.TransientModel):
    _name = 'fms.analysis'
    _description = 'FMS Analysis reports'

    def _default_station_id(self):
        employee = self.env.user.employee_id
        if employee.station_ids:
            return employee.station_ids[0]
        station = self.env['station.station'].search([], limit=1)
        return station

    date_from = fields.Date(string='Date From', required=True)
    date_to = fields.Date(string='Date To', default=fields.Date.today())
    station_id = fields.Many2one(
        'station.station', string='Station', default=_default_station_id)
    report_type = fields.Selection(string='Report Type', 
                                   selection=[
                                       ('wet_summary', 'Wet Summary'),
                                       ('cash_summary', 'Cash Summary'),
                                       ('credit_summary', 'Credit Summary'),
                                       ('daily_report', 'Daily Report')
                                       ],
                                   default='wet_summary')
    
    def _report_mappings(self, report_type):
        return {
            'wet_summary': self._make_wet_summary_report,
            'cash_summary': self._make_cash_summary_report,
            'credit_summary': self._make_credit_summary_report,
            'daily_report':self._make_daily_report,
        }[report_type]
    
    def _prepare_wet_summary_data(self):
        values = {}
        query = f"""
            select 
                sst.sequence, 
                s.date, 
                stst.shift_id, 
                st.name as tank, 
                st.max_volume, 
                stst.opening_qty,
                stst.received_qty,
                stst.sales_qty, 
                stst.closing_dip_qty
            from shift_tank_stock_take stst
            join station_shift s on stst.shift_id = s.id 
            join station_shift_type sst on sst.id = s.type_id 
            join station_tank st on st.id = stst.tank_id
            where 
                s.date >= '{self.date_from}' and
                s.date <= '{self.date_to}' and
                s.station_id = {self.station_id.id} and
                s.state not in {('cancelled', 'draft')}
            order by s.date, sst.sequence;
        """
        self.env.cr.execute(query)
        dippings = self.env.cr.dictfetchall()
        
        for dip in dippings:
            shift_date = dip.get('date')
            tank = dip.get('tank')
            val = {
                    'Date': shift_date,
                    'Opening Stock': dip.get('opening_qty') or 0,
                    'Deliveries': dip.get('received_qty') or 0,
                    'Sales': dip.get('sales_qty') or 0,
                    'Book Stock': lambda row: f'=B{row}+C{row}-D{row}',
                    'Closing Stock': dip.get('closing_dip_qty') or 0,
                    'Variance Loss/Gain': lambda row: f'=F{row}-E{row}',
                    'Cumulative Variance': lambda row: f'=IFERROR(H{row-1}+G{row}, G{row})',
                    'Cumulative Sales': lambda row: f'=IFERROR(I{row-1}+D{row}, D{row})',
                    'Cumulative Percentage (%)': lambda row: f'=IFERROR(round(100 * H{row}/I{row}, 1), 0)'
                }
            if values.get(tank):
                if values[tank].get(shift_date):
                    values[tank][shift_date]['Deliveries'] += dip.get('received_qty') or 0
                    values[tank][shift_date]['Sales'] += dip.get('sales_qty') or 0
                    values[tank][shift_date]['Closing Stock'] += dip.get('closing_dip_qty') or 0 
                else:
                    values[tank][shift_date] = val
            else:
                values[tank] = {shift_date: val, 'volume': dip.get('max_volume')}

        return values

    def _make_wet_summary_report(self, wb):
        ws = wb.active
        fr = 1
        filename = 'Wet Stock Summary Report'
        tanked_values = self._prepare_wet_summary_data()
        for tank, tank_vals in tanked_values.items():
            ws[f'A{fr}'] = f'Tank: {tank}'
            ws.merge_cells(f'A{fr}:F{fr}')
            ws[f'A{fr}'].font = MID_FONT
            
            ws[f'G{fr}'] = f"Capacity: {tank_vals.pop('volume', 0)}"
            ws.merge_cells(f'G{fr}:J{fr}')
            ws[f'G{fr}'].font = MID_FONT
            
            fr += 1
            # Insert columns (after each tank, add columns)
            for index, col in enumerate(WET_SUMMARY_COLS, start=1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = col
                ws[f'{letter}{fr}'].font = MID_FONT
            fr += 1
            
            start_fr = end_fr = fr
            
            # populate tank values
            for val in tank_vals.values():
                for index, col in enumerate(WET_SUMMARY_COLS, start=1):
                    letter = get_column_letter(index)
                    if callable(val.get(col)):
                        ws[f'{letter}{fr}'] = val.get(col).__call__(fr)
                    else:
                        ws[f'{letter}{fr}'] = val.get(col)
                    ws[f'{letter}{fr}'].font = NORMAL_FONT
                fr += 1
            end_fr = fr -1
            
            # populate totals per tank
            ws[f'A{fr}'] = 'TOTALS'
            ws[f'A{fr}'].font = MID_FONT
            for index, col in enumerate(WET_SUMMARY_COLS[1:], start=2):
                    letter = get_column_letter(index)
                    ws[f'{letter}{fr}'] = f"=SUM({letter}{start_fr}:{letter}{end_fr})"
                    ws[f'{letter}{fr}'].font = MID_FONT
            fr += 2
        return wb, filename

    def _prepare_cash_summary_report(self):
        data = defaultdict(dict)
        shifts = self.env['station.shift'].search(
            [('date', '>=', self.date_from), 
             ('date', '<=', self.date_to),
             ('station_id', '=', self.station_id.id)], order='date asc')
        for shift in shifts:
            gun_pms = shift.gun_sale_line.filtered(lambda g: g.tank_id.product_id.default_code == 'PMS')
            gun_ago = shift.gun_sale_line.filtered(lambda g: g.tank_id.product_id.default_code == 'AGO')
            gun_bik = shift.gun_sale_line.filtered(lambda g: g.tank_id.product_id.default_code == 'BIK')
            
            direct_bik = shift.direct_sale_line.filtered(lambda g: g.tank_id.product_id.default_code == 'BIK')
            direct_ago = shift.direct_sale_line.filtered(lambda g: g.tank_id.product_id.default_code == 'AGO')
            direct_pms = shift.direct_sale_line.filtered(lambda g: g.tank_id.product_id.default_code == 'PMS')
            
            pms_qty = sum(gun_pms.mapped('net_sales')) + sum(direct_pms.mapped('quantity'))
            pms_amount = sum(gun_pms.mapped('amount')) + sum(direct_pms.mapped('amount'))
            bik_qty = sum(gun_bik.mapped('net_sales')) + sum(direct_bik.mapped('quantity'))
            bik_amount = sum(gun_bik.mapped('amount')) + sum(direct_bik.mapped('amount'))
            ago_qty = sum(gun_ago.mapped('net_sales')) + sum(direct_ago.mapped('quantity'))
            ago_amount = sum(gun_ago.mapped('amount')) + sum(direct_ago.mapped('amount'))
            
            lubes = sum(shift.dry_sale_line.filtered(lambda d: d.product_id.stock_type == 'lube').mapped('amount'))
            lpg = sum(shift.dry_sale_line.filtered(lambda d: d.product_id.stock_type == 'lpg').mapped('amount'))
            others = sum(shift.other_sale_line.filtered(lambda d: d.product_id.stock_type == 'other').mapped('amount'))
            
            vals = {
                'DATE': shift.date.strftime('%d-%m-%Y'),
                'PMS': pms_qty,
                'PMS AMOUNT': pms_amount, 
                'AGO': ago_qty,
                'AGO AMOUNT': ago_amount,
                'BIK': bik_qty,
                'BIK AMOUNT': bik_amount,
                'LTRS': lambda row: f'=B{row}+D{row}+F{row}',
                'TOTAL AMNT': lambda row: f'=C{row}+E{row}+G{row}',
                'LUBES': lubes,
                'LPG SALES': lpg, 
                'OTHERS': others,
                'RECEIPTS': sum(shift.collection_line.mapped('amount')),
                'TOTAL INCOME': lambda row: f'=SUM(I{row}:M{row})', 
                'CREDIT SALES': sum(shift.credit_sale_line.mapped('amount')),
                'RTT': sum(shift.gun_sale_line.mapped('rtt')),
                'OTHER EXP.': sum(shift.expense_line.mapped('amount')),
                'PAYMENTS': sum(shift.payment_line.mapped('amount')),
                'EXP.BANKING': lambda row: f'=N{row}-O{row}-P{row}-Q{row}',
                'ACT BANKING': shift.cash_banked,
                'DIFF': lambda row: f'=S{row}-R{row}'
            }
            if data.get(shift.date):
                for key, value in data.items():
                    if callable(value) or key == 'DATE':
                        continue
                    data[shift.date][key] += value
            else:
                data[shift.date] = vals
        return list(data.values())
        
    def _make_cash_summary_report(self, wb):
        ws = wb.active
        fr = 1
        filename = 'Cash Summary Report'
        for index, col in enumerate(CASH_SUMMARY_COLS, start=1):
            letter = get_column_letter(index)
            ws[f'{letter}{fr}'] = col
            ws[f'{letter}{fr}'].font = MID_FONT
        fr += 1
        start_fr = fr
        for val in self._prepare_cash_summary_report():
            for index, col in enumerate(CASH_SUMMARY_COLS, start=1):
                letter = get_column_letter(index)
                if callable(val[col]):
                    ws[f'{letter}{fr}'] = val[col].__call__(fr)
                else:
                    ws[f'{letter}{fr}'] = val[col]
                ws[f'{letter}{fr}'].font = NORMAL_FONT
            fr += 1
        end_fr = fr -1
        
        ws[f'A{fr}'] = 'TOTALS'
        ws[f'A{fr}'].font = MID_FONT
        for index, col in enumerate(CASH_SUMMARY_COLS[1:], start=2):
            letter = get_column_letter(index)
            ws[f'{letter}{fr}'] = f"=SUM({letter}{start_fr}:{letter}{end_fr})"
            ws[f'{letter}{fr}'].font = MID_FONT
        return wb, filename

    def _prepare_credit_summary_data(self):
        credit_lines = self.env['shift.credit.sale.line'].search(
            [('shift_id.date', '>=', self.date_from), ('shift_id.date', '<=', self.date_to)],
            order='date asc, station_id')
        data = []
        for line in credit_lines:
            data.append({
                'Date': line.shift_id.date,
                'Station': line.station_id.name,
                'LPO': line.lpo_number,
                'Vehicle No.': line.vehicle_no,
                'Invoice': line.invoice_no,
                'Account Number': line.partner_ref or '',
                'Account Name': line.partner_id.name,
                'Product': line.product_id.name, 
                'LTRS/PCS': line.quantity,
                'Customer Rate': line.price_unit,
                'Amount': line.amount
            })
        return data
  
    def _make_credit_summary_report(self, wb):
        ws = wb.active
        fr = 1
        filename = 'Credit Summary Report'
        for index, col in enumerate(CREDIT_SUMMARY_COLS, start=1):
            letter = get_column_letter(index)
            ws[f'{letter}{fr}'] = col
            ws[f'{letter}{fr}'].font = MID_FONT
        fr += 1
        for val in self._prepare_credit_summary_data():
            for index, col in enumerate(CREDIT_SUMMARY_COLS, start=1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = val.get(col)
                ws[f'{letter}{fr}'].font = NORMAL_FONT
            fr += 1
        return wb, filename
    
    def _prepare_daily_sales_stock_report(self):
        query = """
            SELECT 
                sst.sequence, 
                s.date, 
                stst.shift_id, 
                st.name AS tank, 
                st.max_volume, 
                stst.opening_qty,
                stst.received_qty,
                stst.sales_qty, 
                stst.closing_dip_qty,
                scl.vehicle_no  -- Fetch vehicle from shift credit sale line
            FROM shift_tank_stock_take stst
            JOIN station_shift s ON stst.shift_id = s.id 
            JOIN station_shift_type sst ON sst.id = s.type_id 
            JOIN station_tank st ON st.id = stst.tank_id
            LEFT JOIN shift_credit_sale_line scl ON scl.shift_id = s.id  -- Join to get vehicle
            WHERE 
                s.date >= %s AND
                s.date <= %s AND
                s.station_id = %s AND
                s.state NOT IN ('cancelled', 'draft')
            ORDER BY s.date, sst.sequence;
        """

        # Execute query safely with parameterized inputs
        self.env.cr.execute(query, (self.date_from, self.date_to, self.station_id.id if self.station_id else None))
        sales_data = self.env.cr.dictfetchall()

        # Group data
        report_data = []
        for record in sales_data:
            report_data.append({
                'Date': record.get('date').strftime('%d/%m/%Y') if record.get('date') else '',
                'Shift ID': record.get('shift_id'),
                'Tank': record.get('tank'),
                'Max Volume': record.get('max_volume') or 0,
                'Opening Quantity': record.get('opening_qty') or 0,
                'Received Quantity': record.get('received_qty') or 0,
                'Sales Quantity': record.get('sales_qty') or 0,
                'Closing Dip Quantity': record.get('closing_dip_qty') or 0,
                'Vehicle': record.get('vehicle_no') or 'N/A'  # Use vehicle from shift credit sale line
            })

        return report_data
    
    def _make_daily_report(self, wb):
        """
        Generate the daily sales stock report in Excel format.
        
        Args:
            wb: xlsxwriter.Workbook instance to write data into.
        
        Returns:
            wb, filename (Workbook and filename for saving).
        """

        # Fetch all sales orders
        records = self.env['sale.order'].search([])

        # Create a worksheet
        sheet = wb.add_worksheet("Daily Sales Report")

        # Write column headers
        headers = ["Customer", "Invoice No", "Vehicle", "Product", "Quantity"]
        for col, header in enumerate(headers):
            sheet.write(0, col, header)

        row = 1  # Start from the second row (0-based index)

        for order in records:
            customer = order.partner_id.name or "Unknown"
            invoice_no = order.invoice_ids and order.invoice_ids[0].name or "N/A"
            vehicle = order.vehicle if hasattr(order, 'vehicle') else "N/A"

            # Fetch sale order lines
            for line in order.order_line:
                product = line.product_id.name or "Unknown"
                qty = line.product_uom_qty or 0.0

                sheet.write(row, 0, customer)
                sheet.write(row, 1, invoice_no)
                sheet.write(row, 2, vehicle)
                sheet.write(row, 3, product)
                sheet.write(row, 4, qty)

                row += 1

        return wb, "daily_sales_report.xlsx"

    def action_generate_report(self):
        """ Generate the report, store it as an attachment, and provide a download link. """
        
        output = BytesIO()  # Create an in-memory buffer
        wb = xlsxwriter.Workbook(output)  # Initialize workbook
        
        # Fetch the appropriate report function
        report_func = self._report_mappings(self.report_type)
        
        if callable(report_func):
            wb, filename = report_func(wb)  # Generate the report

        # Close the workbook to save content
        wb.close()
        
        # Prepare the file for download
        output.seek(0)
        excel_data = output.getvalue()
        output.close()

        # Store file as an attachment
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(excel_data),
            'res_model': 'fms.analysis',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Provide download link
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
class SaleOrder(models.Model):
    _inherit = 'sale.order'

    vehicle = fields.Char(string="Vehicle")
class ReportDownload(models.TransientModel):
    _name = 'excel.wizard'
    _description = 'Download Excel Forms'

    name = fields.Char('File Name', size=64)
    report = fields.Binary('Your Report', readonly=True)

    @api.model
    def create_xls(self):
        _, xls_path = tempfile.mkstemp(
            suffix='.xlsx', prefix='xlsreport.tmp.')
        return xls_path

    @api.model
    def save_xls_file(self, xls_path, message=None):
        if not message:
            message = f'A report has been generated by {self.env.user.name}'
        with open(xls_path, 'rb') as f:
            datas = base64.encodebytes(f.read())
            _logger.info(message)
            self.delete_tempfile(xls_path)
        return datas

    @api.model
    def delete_tempfile(self, path):
        try:
            os.unlink(path)
        except (OSError, IOError):
            _logger.error('Error when trying to remove file %s' % path)
